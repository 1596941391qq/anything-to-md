"""
Core converter module - wraps MarkItDown with additional capabilities
"""

import os
import shutil
import subprocess
import json
import uuid
from pathlib import Path
from typing import Optional, List, Callable
from dataclasses import dataclass, field

# Lazy import to avoid import errors when markitdown is not installed
_markitdown = None

def _get_markitdown():
    global _markitdown
    if _markitdown is None:
        try:
            from markitdown import MarkItDown
            _markitdown = MarkItDown
        except ImportError:
            raise ImportError(
                "markitdown is required. Install it with: pip install markitdown"
            )
    return _markitdown

from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
import pathspec

console = Console()


@dataclass
class ConversionResult:
    """Result of a single file conversion"""
    source_path: Path
    target_path: Path
    success: bool
    markdown: Optional[str] = None
    error: Optional[str] = None
    skipped: bool = False
    skip_reason: Optional[str] = None


@dataclass
class BatchResult:
    """Result of a batch conversion"""
    source_dir: Path
    target_dir: Path
    total_files: int = 0
    converted: int = 0
    skipped: int = 0
    failed: int = 0
    results: List[ConversionResult] = field(default_factory=list)
    
    @property
    def success_rate(self) -> float:
        if self.total_files == 0:
            return 0.0
        return self.converted / self.total_files * 100


class AnythingToMD:
    """
    Universal document to Markdown converter.
    
    Supports:
    - PDF, Word, Excel, PowerPoint
    - Images (with OCR)
    - Audio/Video (transcription)
    - YouTube URLs
    - HTML, EPUB, and more
    """
    
    # File types that should be skipped
    SKIP_EXTENSIONS = {
        # Already text/markdown
        '.md', '.txt', '.json', '.xml', '.yaml', '.yml',
        # Code files (optional - can be converted to MD with syntax highlighting)
        '.py', '.js', '.ts', '.java', '.c', '.cpp', '.h', '.hpp',
        '.go', '.rs', '.rb', '.php', '.swift', '.kt',
        # Config files
        '.toml', '.ini', '.cfg', '.conf',
        # Lock files
        '.lock', '.sum',
        # Binary non-convertible
        '.exe', '.dll', '.so', '.dylib', '.bin',
        '.zip', '.tar', '.gz', '.rar', '.7z',  # archives are handled by markitdown
        # Database files
        '.db', '.sqlite', '.sqlite3',
    }
    
    # Extensions that are already Markdown-friendly
    KEEP_AS_IS_EXTENSIONS = {
        '.md', '.markdown', '.rst'
    }
    
    def __init__(
        self,
        enable_plugins: bool = False,
        skip_patterns: Optional[List[str]] = None,
        verbose: bool = True
    ):
        self.enable_plugins = enable_plugins
        self.markitdown = None
        self.markitdown_error = None
        self.verbose = verbose
        
        # Default skip patterns (gitignore style)
        default_skips = [
            '.git/',
            '__pycache__/',
            'node_modules/',
            '.venv/',
            'venv/',
            '.env',
            '*.pyc',
            '*.pyo',
            '.DS_Store',
            'Thumbs.db',
        ]
        self.skip_patterns = default_skips + (skip_patterns or [])
        self.spec = pathspec.PathSpec.from_lines(
            pathspec.patterns.GitWildMatchPattern,
            self.skip_patterns
        )

        # Best effort: if MarkItDown import/runtime is broken, keep fallback mode available.
        self._ensure_markitdown()

    def _ensure_markitdown(self):
        if self.markitdown is not None:
            return self.markitdown
        if self.markitdown_error is not None:
            return None

        try:
            MarkItDown = _get_markitdown()
            self.markitdown = MarkItDown(enable_plugins=self.enable_plugins)
            return self.markitdown
        except Exception as ex:
            self.markitdown_error = str(ex)
            if self.verbose:
                console.print(
                    f"[yellow]MarkItDown unavailable, switching to fallback mode:[/yellow] {ex}"
                )
            return None
    
    def should_skip(self, path: Path, base_dir: Path) -> tuple[bool, str]:
        """Check if a file should be skipped"""
        rel_path = path.relative_to(base_dir)
        rel_str = str(rel_path).replace('\\', '/')
        
        # Check gitignore patterns
        if self.spec.match_file(rel_str):
            return True, "Matched skip pattern"
        
        # Check extension
        ext = path.suffix.lower()
        if ext in self.SKIP_EXTENSIONS:
            return True, f"Extension {ext} in skip list"
        
        return False, ""
    
    def _is_media_file(self, path: Path) -> bool:
        media_exts = {
            '.mp3', '.wav', '.m4a', '.flac', '.ogg', '.aac',
            '.mp4', '.mkv', '.avi', '.mov', '.webm', '.wmv', '.m4v'
        }
        return path.suffix.lower() in media_exts

    def _is_video_file(self, path: Path) -> bool:
        video_exts = {'.mp4', '.mkv', '.avi', '.mov', '.webm', '.wmv', '.m4v'}
        return path.suffix.lower() in video_exts

    def _parse_subtitle_cues(self, subtitle_text: str) -> List[tuple[str, str]]:
        """Parse SRT/WebVTT-like subtitle text into (timestamp, content) cues."""
        text = subtitle_text.replace('\r\n', '\n').replace('\r', '\n').strip()
        if not text:
            return []

        cues: List[tuple[str, str]] = []
        for block in text.split('\n\n'):
            lines = [line.strip() for line in block.split('\n') if line.strip()]
            if not lines:
                continue

            time_idx = -1
            for i, line in enumerate(lines):
                if '-->' in line:
                    time_idx = i
                    break
            if time_idx < 0:
                continue

            start = lines[time_idx].split('-->', 1)[0].strip()
            content = ' '.join(lines[time_idx + 1:]).strip()
            if content:
                cues.append((start, content))
        return cues

    def _subtitle_markdown_section(self, section_title: str, subtitle_text: str) -> List[str]:
        cues = self._parse_subtitle_cues(subtitle_text)
        lines = [f'### {section_title}', '']
        if cues:
            lines.extend([f'- [{timestamp}] {content}' for timestamp, content in cues])
        elif subtitle_text.strip():
            lines.extend(['```text', subtitle_text.strip(), '```'])
        else:
            lines.append('- Empty subtitle content')
        lines.append('')
        return lines

    def _get_video_subtitle_streams(self, source: Path) -> List[dict]:
        try:
            proc = subprocess.run(
                ['ffprobe', '-v', 'error', '-show_streams', '-print_format', 'json', str(source)],
                capture_output=True,
                text=True,
                timeout=30,
                check=False,
            )
            if proc.returncode != 0 or not proc.stdout.strip():
                return []
            data = json.loads(proc.stdout)
            streams = data.get('streams', [])
            return [s for s in streams if s.get('codec_type') == 'subtitle']
        except Exception:
            return []

    def _extract_embedded_subtitle_text(self, source: Path, stream_index: int) -> Optional[str]:
        try:
            proc = subprocess.run(
                [
                    'ffmpeg', '-v', 'error', '-i', str(source),
                    '-map', f'0:{stream_index}',
                    '-f', 'srt', '-'
                ],
                capture_output=True,
                text=True,
                timeout=60,
                check=False,
            )
            if proc.returncode != 0:
                return None
            text = proc.stdout.strip()
            return text if text else None
        except Exception:
            return None

    def _extract_video_subtitles_markdown(self, source: Path) -> List[str]:
        if not self._is_video_file(source):
            return []

        lines: List[str] = []

        # Use sidecar subtitle files first when present.
        sidecar_exts = ('.srt', '.vtt')
        for ext in sidecar_exts:
            sidecar = source.with_suffix(ext)
            if not sidecar.exists():
                continue
            try:
                text = sidecar.read_text(encoding='utf-8', errors='replace')
                lines.extend(self._subtitle_markdown_section(f'Sidecar Subtitle: {sidecar.name}', text))
            except Exception as ex:
                lines.extend([
                    f'### Sidecar Subtitle: {sidecar.name}',
                    '',
                    f'- Failed to read subtitle file: `{ex}`',
                    '',
                ])

        # Then extract embedded subtitle tracks from the video container.
        for stream in self._get_video_subtitle_streams(source):
            idx = stream.get('index')
            if idx is None:
                continue

            codec = stream.get('codec_name', 'unknown')
            language = stream.get('tags', {}).get('language', 'und')
            title = f'Embedded Subtitle Stream #{idx} ({codec}, lang={language})'

            text = self._extract_embedded_subtitle_text(source, idx)
            if text:
                lines.extend(self._subtitle_markdown_section(title, text))
            else:
                lines.extend([f'### {title}', '', '- Subtitle stream exists but extraction returned empty.', ''])

        return lines

    def _format_seconds(self, seconds: int) -> str:
        h = seconds // 3600
        m = (seconds % 3600) // 60
        s = seconds % 60
        return f'{h:02d}:{m:02d}:{s:02d}'

    def _transcribe_media_audio_markdown(self, source: Path, segment_seconds: int = 45) -> List[str]:
        """
        Best-effort transcription fallback for audio/video files.
        Uses ffmpeg for segmentation and SpeechRecognition (Google recognizer).
        """
        if not self._is_media_file(source):
            return []

        try:
            import speech_recognition as sr
        except Exception:
            return []

        import tempfile
        temp_root = Path(tempfile.gettempdir()) / 'anything-to-md-transcribe'
        temp_root.mkdir(parents=True, exist_ok=True)
        run_id = uuid.uuid4().hex
        pattern = temp_root / f'seg_{run_id}_%04d.wav'

        try:
            # Segmenting avoids single huge recognition requests and improves accuracy.
            seg_proc = subprocess.run(
                [
                    'ffmpeg', '-y', '-v', 'error',
                    '-i', str(source),
                    '-vn',
                    '-ac', '1',
                    '-ar', '16000',
                    '-f', 'segment',
                    '-segment_time', str(segment_seconds),
                    '-reset_timestamps', '1',
                    str(pattern),
                ],
                capture_output=True,
                text=True,
                timeout=600,
                check=False,
            )
            if seg_proc.returncode != 0:
                return []

            segment_files = sorted(temp_root.glob(f'seg_{run_id}_*.wav'))
            if not segment_files:
                return []

            recognizer = sr.Recognizer()
            transcript_lines = []
            for idx, seg in enumerate(segment_files):
                with sr.AudioFile(str(seg)) as audio_source:
                    audio_data = recognizer.record(audio_source)
                try:
                    text = recognizer.recognize_google(audio_data)
                    text = text.strip()
                except Exception:
                    text = ''

                if text:
                    ts = self._format_seconds(idx * segment_seconds)
                    transcript_lines.append(f'- [{ts}] {text}')

            if not transcript_lines:
                return []

            return ['## Extracted Transcript (Audio Fallback)', ''] + transcript_lines + ['']
        except Exception:
            return []
        finally:
            for seg in temp_root.glob(f'seg_{run_id}_*.wav'):
                try:
                    seg.unlink()
                except Exception:
                    pass

    def _extract_pdf_markdown_fallback(self, source: Path) -> List[str]:
        if source.suffix.lower() != '.pdf':
            return []

        try:
            from pypdf import PdfReader
        except Exception:
            return []

        try:
            reader = PdfReader(str(source))
            lines: List[str] = ['## Extracted PDF Text (Fallback)', '']
            non_empty_pages = 0

            for page_idx, page in enumerate(reader.pages, start=1):
                text = page.extract_text() or ''
                text = text.strip()
                if not text:
                    continue
                non_empty_pages += 1
                lines.append(f'### Page {page_idx}')
                lines.append('')
                lines.append(text)
                lines.append('')

            if non_empty_pages == 0:
                return []
            return lines
        except Exception:
            return []

    def _build_fallback_markdown(self, source: Path, error: str) -> str:
        lines = [
            f'# Conversion Fallback: {source.name}',
            '',
            f'- Source: `{source}`',
            f'- Reason: `{error}`',
            '',
        ]

        if self._is_video_file(source):
            subtitle_lines = self._extract_video_subtitles_markdown(source)
            if subtitle_lines:
                lines += ['## Extracted Subtitles', ''] + subtitle_lines
            transcript_lines = self._transcribe_media_audio_markdown(source)
            if transcript_lines:
                lines += transcript_lines

        # xlsx fallback using openpyxl when MarkItDown dependencies are broken.
        if source.suffix.lower() in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
            try:
                from openpyxl import load_workbook

                wb = load_workbook(source, data_only=True, read_only=True)
                lines += ['## Workbook Preview', '']
                for ws in wb.worksheets:
                    lines += [f'### Sheet: {ws.title}', '']
                    row_count = 0
                    for row in ws.iter_rows(values_only=True):
                        vals = [str(v).strip() for v in row if v is not None and str(v).strip() != '']
                        if vals:
                            lines.append('- ' + ' | '.join(vals))
                        row_count += 1
                        if row_count >= 50:
                            lines.append('- ... (truncated)')
                            break
                    lines.append('')
            except Exception as ex:
                lines += [f'Workbook fallback failed: `{ex}`', '']

        if source.suffix.lower() == '.pdf':
            lines += self._extract_pdf_markdown_fallback(source)

        if len(lines) <= 5:
            # Only metadata header, no real content extracted — treat as failure
            return None

        return '\n'.join(lines).strip() + '\n'

    def convert_file(
        self,
        source: Path,
        output_dir: Optional[Path] = None,
        output_name: Optional[str] = None
    ) -> ConversionResult:
        """Convert a single file to Markdown"""
        
        if not source.exists():
            return ConversionResult(
                source_path=source,
                target_path=Path(""),
                success=False,
                error=f"Source file not found: {source}"
            )
        
        # Determine output path
        if output_dir is None:
            output_dir = source.parent
        
        if output_name is None:
            output_name = source.stem + '.md'
        elif not output_name.endswith('.md'):
            output_name = output_name + '.md'
        
        target_path = output_dir / output_name
        
        converter = self._ensure_markitdown()

        try:
            # Use MarkItDown to convert
            if converter is None:
                raise RuntimeError(self.markitdown_error or "MarkItDown unavailable")
            result = converter.convert(str(source))
            # Handle different API versions - result may be string or object
            if hasattr(result, 'text_content'):
                markdown_content = result.text_content
            elif hasattr(result, 'markdown'):
                markdown_content = result.markdown
            else:
                markdown_content = str(result)
            
            # Write output
            output_dir.mkdir(parents=True, exist_ok=True)
            target_path.write_text(markdown_content, encoding='utf-8')
            
            if self.verbose:
                console.print(f"[green]✓[/green] {source.name} -> {output_name}")
            
            return ConversionResult(
                source_path=source,
                target_path=target_path,
                success=True,
                markdown=markdown_content
            )
            
        except Exception as e:
            error_text = str(e)
            fallback_md = self._build_fallback_markdown(source, error_text)

            if fallback_md is None:
                if self.verbose:
                    console.print(f"[red]✗[/red] {source.name}: no content extracted (metadata only)")
                return ConversionResult(
                    source_path=source,
                    target_path=target_path,
                    success=False,
                    error=error_text
                )

            try:
                output_dir.mkdir(parents=True, exist_ok=True)
                target_path.write_text(fallback_md, encoding='utf-8')
                if self.verbose:
                    console.print(f"[yellow]![/yellow] {source.name}: fallback markdown generated")

                return ConversionResult(
                    source_path=source,
                    target_path=target_path,
                    success=True,
                    markdown=fallback_md,
                    error=error_text
                )
            except Exception as write_err:
                if self.verbose:
                    console.print(f"[red]✗[/red] {source.name}: {write_err}")

                return ConversionResult(
                    source_path=source,
                    target_path=target_path,
                    success=False,
                    error=f"{error_text}; fallback write failed: {write_err}"
                )

    def convert_directory(
        self,
        source_dir: Path,
        target_dir: Path,
        preserve_structure: bool = True,
        copy_non_convertible: bool = False,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> BatchResult:
        """
        Convert all files in a directory to Markdown.
        
        Args:
            source_dir: Source directory to process
            target_dir: Target directory for output
            preserve_structure: Keep the original directory structure
            copy_non_convertible: Copy files that can't be converted
            progress_callback: Optional callback(current, total, filename)
        
        Returns:
            BatchResult with conversion statistics
        """
        source_dir = Path(source_dir).resolve()
        target_dir = Path(target_dir).resolve()
        
        if not source_dir.exists():
            raise ValueError(f"Source directory not found: {source_dir}")

        target_inside_source = source_dir == target_dir or source_dir in target_dir.parents

        # Collect all files
        all_files = [f for f in source_dir.rglob('*') if f.is_file()]
        convertible_files = []
        skip_results = []
        
        for f in all_files:
            if target_inside_source and (target_dir == f.parent or target_dir in f.parents):
                skip_results.append(ConversionResult(
                    source_path=f,
                    target_path=Path(""),
                    success=False,
                    skipped=True,
                    skip_reason="Output directory excluded from scan"
                ))
                continue

            should_skip, reason = self.should_skip(f, source_dir)
            if should_skip:
                skip_results.append(ConversionResult(
                    source_path=f,
                    target_path=Path(""),
                    success=False,
                    skipped=True,
                    skip_reason=reason
                ))
            else:
                convertible_files.append(f)
        
        result = BatchResult(
            source_dir=source_dir,
            target_dir=target_dir,
            total_files=len(convertible_files),
            skipped=len(skip_results),
            results=skip_results
        )
        
        if not convertible_files:
            if self.verbose:
                console.print("[yellow]No convertible files found[/yellow]")
            return result
        
        if self.verbose:
            console.print(f"[cyan]Processing {len(convertible_files)} files...[/cyan]")
        
        # Process files with progress
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
            disable=not self.verbose
        ) as progress:
            task = progress.add_task("Converting...", total=len(convertible_files))
            
            for i, file_path in enumerate(convertible_files):
                progress.update(task, description=f"Converting {file_path.name}")
                
                # Calculate relative path and target
                rel_path = file_path.relative_to(source_dir)
                
                if preserve_structure:
                    # Keep directory structure
                    target_subdir = target_dir / rel_path.parent
                    output_name = file_path.name + '.md'
                else:
                    # Flat structure
                    target_subdir = target_dir
                    # Add parent dir prefix to avoid name collisions
                    if rel_path.parent != Path('.'):
                        output_name = f"{rel_path.parent}_{file_path.name}".replace('/', '_').replace('\\', '_') + '.md'
                    else:
                        output_name = file_path.name + '.md'
                
                target_subdir.mkdir(parents=True, exist_ok=True)
                
                conv_result = self.convert_file(
                    source=file_path,
                    output_dir=target_subdir,
                    output_name=output_name
                )
                
                result.results.append(conv_result)
                
                if conv_result.success:
                    result.converted += 1
                else:
                    result.failed += 1
                
                progress.advance(task)
                
                if progress_callback:
                    progress_callback(i + 1, len(convertible_files), file_path.name)
        
        if self.verbose:
            console.print(
                f"\n[bold]Results:[/bold] "
                f"[green]{result.converted} converted[/green], "
                f"[yellow]{result.skipped} skipped[/yellow], "
                f"[red]{result.failed} failed[/red]"
            )
        
        return result
    
    def convert_youtube(
        self,
        url: str,
        output_path: Optional[Path] = None
    ) -> ConversionResult:
        """Convert YouTube video to Markdown transcript"""
        
        try:
            converter = self._ensure_markitdown()
            if converter is None:
                raise RuntimeError(self.markitdown_error or "MarkItDown unavailable")
            result = converter.convert(url)
            # Handle different API versions
            if hasattr(result, 'text_content'):
                content = result.text_content
            elif hasattr(result, 'markdown'):
                content = result.markdown
            else:
                content = str(result)
            
            if output_path:
                output_path = Path(output_path)
                output_path.parent.mkdir(parents=True, exist_ok=True)
                output_path.write_text(content, encoding='utf-8')
            
            return ConversionResult(
                source_path=Path(url),
                target_path=output_path or Path(""),
                success=True,
                markdown=content
            )
            
        except Exception as e:
            return ConversionResult(
                source_path=Path(url),
                target_path=Path(""),
                success=False,
                error=str(e)
            )
